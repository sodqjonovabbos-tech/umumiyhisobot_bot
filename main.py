import asyncio
import calendar
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path

from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from sqlalchemy import select, func
from openpyxl import Workbook

from config import BOT_TOKEN, OWNER_ID

TZ = ZoneInfo("Asia/Tashkent")
from app.database.db import init_db, SessionLocal
from app.database.models import Income, Expense, Debt, DebtPayment

router = Router()

EXPENSE_CATEGORIES = [
    "Soliq", "Kommunal to'lov", "Onamga", "Xotinimga", "Sovg'a va ehson",
    "Odamlarga qarzga berdim", "O'qishga", "Sayohatga", "Shtraflarga",
    "Ovqat ko'cha", "Uyga", "Bolalar uchun", "Maktabga",
    "Mashinaga", "Tualetniy bumaga", "Boshqa xarajat"
]
INCOME_SOURCES = ["Oylik", "Biznes", "Qaytgan qarz", "Sovg'a pul", "Boshqa kirim"]

class AddIncome(StatesGroup):
    source = State()
    amount = State()
    note = State()

class AddExpense(StatesGroup):
    category = State()
    amount = State()
    note = State()

class AddDebt(StatesGroup):
    debt_type = State()
    person = State()
    name = State()
    amount = State()
    note = State()

class PayDebt(StatesGroup):
    person = State()
    amount = State()

class SearchState(StatesGroup):
    query = State()


def money(value: int) -> str:
    return f"{int(value):,}".replace(',', ' ') + " so'm"


def parse_amount(text: str) -> int | None:
    cleaned = text.replace(' ', '').replace(',', '').replace('.', '').strip()
    if not cleaned.isdigit():
        return None
    return int(cleaned)


def owner_only(message: Message) -> bool:
    return OWNER_ID and message.from_user and message.from_user.id == OWNER_ID


def now_tashkent() -> datetime:
    return datetime.now(TZ).replace(tzinfo=None)


def main_menu():
    kb = ReplyKeyboardBuilder()
    buttons = [
        "➕ Kirim qo'shish", "➖ Chiqim qo'shish",
        "💳 Qarzlarim", "📊 Oylik hisobot",
        "📤 Excel eksport", "🔍 Qidirish"
    ]
    for b in buttons:
        kb.button(text=b)
    kb.adjust(2)
    return kb.as_markup(resize_keyboard=True)


def inline_list(prefix: str, items: list[str]):
    kb = InlineKeyboardBuilder()
    for item in items:
        kb.button(text=item, callback_data=f"{prefix}:{item}")
    kb.adjust(1)
    return kb.as_markup()


@router.message(CommandStart())
async def start(message: Message):
    if not owner_only(message):
        await message.answer("⛔ Bu bot faqat egasi uchun ishlaydi.")
        return
    await message.answer(
        "Assalomu alaykum!\n\n"
        "Bu bot pul oqimini nazorat qiladi.\n"
        "Kirim, chiqim, qarz va oylik hisobotlarni yozib boring.",
        reply_markup=main_menu()
    )


@router.message(F.text == "➕ Kirim qo'shish")
async def add_income_start(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddIncome.source)
    await message.answer(
        "Kirim manbasini tanlang:\n\nOrqaga qaytish uchun /back yozing.",
        reply_markup=inline_list('income_source', INCOME_SOURCES)
    )


@router.message(AddIncome.source, F.text.casefold() == "/back")
async def income_source_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.clear()
    await message.answer("⬅️ Bosh menyuga qaytdingiz.", reply_markup=main_menu())


@router.callback_query(F.data.startswith('income_source:'))
async def income_source(call: CallbackQuery, state: FSMContext):
    source = call.data.split(':', 1)[1]
    await state.update_data(source=source)
    await state.set_state(AddIncome.amount)
    await call.message.answer("Qancha pul tushdi? Masalan: 5000000\n\nOrqaga qaytish uchun /back yozing.")
    await call.answer()


@router.message(AddIncome.amount, F.text.casefold() == "/back")
async def income_amount_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddIncome.source)
    await message.answer(
        "⬅️ Kirim manbasini qayta tanlang:",
        reply_markup=inline_list('income_source', INCOME_SOURCES)
    )


@router.message(AddIncome.amount)
async def income_amount(message: Message, state: FSMContext):
    amount = parse_amount(message.text or '')
    if amount is None:
        await message.answer("Summani faqat raqam bilan yozing. Masalan: 5000000")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddIncome.note)
    await message.answer("Izoh yozing. Masalan: Oylik tushdi\n\nSummani qayta kiritish uchun /back yozing.")


@router.message(AddIncome.note, F.text.casefold() == "/back")
async def income_note_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddIncome.amount)
    await message.answer("⬅️ Summani qayta kiriting. Masalan: 5000000")


@router.message(AddIncome.note)
async def income_note(message: Message, state: FSMContext):
    data = await state.get_data()
    async with SessionLocal() as session:
        session.add(Income(amount=data['amount'], source=data['source'], note=message.text or ''))
        await session.commit()
    await state.clear()
    await message.answer(
        f"✅ Kirim saqlandi\n{data['source']}: {money(data['amount'])}\nIzoh: {message.text}\n\n"
        "❗ Noto'g'ri kiritilgan bo'lsa, oxirgi yozuvni bekor qilish uchun /back yozing.",
        reply_markup=main_menu()
    )


@router.message(F.text == "➖ Chiqim qo'shish")
async def add_expense_start(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddExpense.category)
    await message.answer(
        "Chiqim kategoriyasini tanlang:\n\nOrqaga qaytish uchun /back yozing.",
        reply_markup=inline_list('expense_cat', EXPENSE_CATEGORIES)
    )


@router.message(AddExpense.category, F.text.casefold() == "/back")
async def expense_category_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.clear()
    await message.answer("⬅️ Bosh menyuga qaytdingiz.", reply_markup=main_menu())


@router.callback_query(F.data.startswith('expense_cat:'))
async def expense_category(call: CallbackQuery, state: FSMContext):
    category = call.data.split(':', 1)[1]
    await state.update_data(category=category)
    await state.set_state(AddExpense.amount)
    await call.message.answer("Qancha pul ishlatildi? Masalan: 10000\n\nOrqaga qaytish uchun /back yozing.")
    await call.answer()


@router.message(AddExpense.amount, F.text.casefold() == "/back")
async def expense_amount_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddExpense.category)
    await message.answer(
        "⬅️ Chiqim kategoriyasini qayta tanlang:",
        reply_markup=inline_list('expense_cat', EXPENSE_CATEGORIES)
    )


@router.message(AddExpense.amount)
async def expense_amount(message: Message, state: FSMContext):
    amount = parse_amount(message.text or '')
    if amount is None:
        await message.answer("Summani faqat raqam bilan yozing. Masalan: 10000")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddExpense.note)
    await message.answer("Izoh yozing. Masalan: Gazga\n\nSummani qayta kiritish uchun /back yozing.")


@router.message(AddExpense.note, F.text.casefold() == "/back")
async def expense_note_back(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(AddExpense.amount)
    await message.answer("⬅️ Summani qayta kiriting. Masalan: 10000")


@router.message(AddExpense.note)
async def expense_note(message: Message, state: FSMContext):
    data = await state.get_data()
    async with SessionLocal() as session:
        session.add(Expense(amount=data['amount'], category=data['category'], note=message.text or ''))
        await session.commit()
    await state.clear()
    await message.answer(
        f"✅ Chiqim saqlandi\n{data['category']}: {money(data['amount'])}\nIzoh: {message.text}\n\n"
        "❗ Noto'g'ri kiritilgan bo'lsa, oxirgi yozuvni bekor qilish uchun /back yozing.",
        reply_markup=main_menu()
    )


@router.message(F.text.casefold() == "/back")
async def undo_last_saved_transaction(message: Message, state: FSMContext):
    """Saqlanib bo'lgan eng oxirgi kirim yoki chiqimni bekor qiladi."""
    if not owner_only(message):
        return

    # Boshqa jarayon ochiq bo'lsa, pul yozuvini o'chirmaymiz — jarayonni bekor qilamiz.
    current_state = await state.get_state()
    if current_state is not None:
        await state.clear()
        await message.answer(
            "⬅️ Joriy amal bekor qilindi. Bosh menyuga qaytdingiz.",
            reply_markup=main_menu()
        )
        return

    async with SessionLocal() as session:
        last_income = (await session.execute(
            select(Income).order_by(Income.created_at.desc(), Income.id.desc()).limit(1)
        )).scalar_one_or_none()
        last_expense = (await session.execute(
            select(Expense).order_by(Expense.created_at.desc(), Expense.id.desc()).limit(1)
        )).scalar_one_or_none()

        if last_income is None and last_expense is None:
            await message.answer(
                "Bekor qilish uchun kirim yoki chiqim topilmadi.",
                reply_markup=main_menu()
            )
            return

        # Ikki jadval ichidan vaqt bo'yicha eng oxirgi saqlangan yozuv tanlanadi.
        undo_income = False
        if last_income is not None and last_expense is None:
            undo_income = True
        elif last_income is not None and last_expense is not None:
            undo_income = last_income.created_at >= last_expense.created_at

        if undo_income:
            amount = last_income.amount
            title = last_income.source
            note = last_income.note
            await session.delete(last_income)
            kind = "kirim"
            icon = "➕"
        else:
            amount = last_expense.amount
            title = last_expense.category
            note = last_expense.note
            await session.delete(last_expense)
            kind = "chiqim"
            icon = "➖"

        await session.commit()

    if undo_income:
        await state.update_data(source=title)
        await state.set_state(AddIncome.amount)
        prompt = "To'g'ri kirim summasini yozing. Masalan: 500000"
    else:
        await state.update_data(category=title)
        await state.set_state(AddExpense.amount)
        prompt = "To'g'ri chiqim summasini yozing. Masalan: 500000"

    await message.answer(
        f"↩️ Oxirgi {kind} bekor qilindi.\n"
        f"{icon} {title}: {money(amount)}\n"
        f"Izoh: {note or '-'}\n\n"
        f"{prompt}\n\n"
        "Boshqa bo'limga qaytish uchun yana /back yozing."
    )


@router.message(F.text == "💳 Qarzlarim")
async def debts_menu(message: Message):
    if not owner_only(message): return
    kb = InlineKeyboardBuilder()
    kb.button(text="📥 Qarz oldim", callback_data="debt_borrow")
    kb.button(text="📤 Qarz berdim", callback_data="debt_lend")
    kb.button(text="📋 Qarzlar ro'yxati", callback_data="debt_list")
    kb.button(text="📥 Qarz Excel yuklash", callback_data="debt_excel")
    kb.adjust(1)
    await message.answer("Qarzlar bo'limi:", reply_markup=kb.as_markup())


@router.callback_query(F.data == 'debt_borrow')
async def debt_add_type(call: CallbackQuery, state: FSMContext):
    if not call.from_user or call.from_user.id != OWNER_ID:
        await call.answer("Bu amal siz uchun ruxsat etilmagan.", show_alert=True)
        return
    await state.update_data(debt_type='oldim')
    await state.set_state(AddDebt.person)
    await call.message.answer("Kimdan qarz oldingiz? Ismini yozing. Masalan: Ali aka")
    await call.answer()


@router.message(AddDebt.person)
async def debt_person(message: Message, state: FSMContext):
    await state.update_data(person=message.text or '')
    await state.set_state(AddDebt.name)
    await message.answer("Qarz nomini yozing. Masalan: Mashina, Uy, O'qish, Naqd pul")


@router.message(AddDebt.name)
async def debt_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(AddDebt.amount)
    await message.answer("Umumiy qarz summasini yozing. Masalan: 100000000")


@router.message(AddDebt.amount)
async def debt_amount(message: Message, state: FSMContext):
    amount = parse_amount(message.text or '')
    if amount is None:
        await message.answer("Summani faqat raqam bilan yozing.")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddDebt.note)
    await message.answer("Nega qarz oldingiz? Izoh yozing. Masalan: Mashina boshlang'ich to'lovi uchun")


@router.message(AddDebt.note)
async def debt_note(message: Message, state: FSMContext):
    data = await state.get_data()
    async with SessionLocal() as session:
        session.add(Debt(
            debt_type=data.get('debt_type', 'oldim'),
            person=data.get('person', ''),
            name=data['name'],
            total_amount=data['amount'],
            note=message.text or ''
        ))
        await session.commit()
    await state.clear()
    turi = "Qarz oldim" if data.get('debt_type') == 'oldim' else "Qarz berdim"
    kim = "Kimdan" if data.get('debt_type') == 'oldim' else "Kimga"
    await message.answer(
        f"✅ {turi} saqlandi\n"
        f"{kim}: {data.get('person', '')}\n"
        f"Nomi: {data['name']}\n"
        f"Summa: {money(data['amount'])}\n"
        f"Izoh: {message.text}",
        reply_markup=main_menu()
    )


async def debt_paid_amount(session, debt_id: int) -> int:
    paid = await session.scalar(
        select(func.coalesce(func.sum(DebtPayment.amount), 0)).where(
            DebtPayment.debt_id == debt_id
        )
    )
    return int(paid or 0)


async def active_borrowed_people() -> list[dict]:
    """Qarzimiz bor odamlarni ism bo'yicha guruhlab qaytaradi."""
    async with SessionLocal() as session:
        debts = (
            await session.execute(
                select(Debt)
                .where(Debt.debt_type == 'oldim')
                .order_by(Debt.created_at.asc(), Debt.id.asc())
            )
        ).scalars().all()

        people: dict[str, dict] = {}
        for debt in debts:
            paid = await debt_paid_amount(session, debt.id)
            remaining = max(0, int(debt.total_amount) - paid)
            if remaining <= 0:
                continue

            person_name = (getattr(debt, 'person', '') or "Noma'lum").strip()
            person_key = person_name.casefold()
            if person_key not in people:
                people[person_key] = {
                    'key': person_key,
                    'person': person_name,
                    'remaining': 0,
                }
            people[person_key]['remaining'] += remaining

        return list(people.values())


async def debts_text():
    async with SessionLocal() as session:
        debts = (await session.execute(select(Debt).order_by(Debt.id.desc()))).scalars().all()
        if not debts:
            return "Faol qarzlar yo'q."
        lines = ["📋 Qarzlar ro'yxati:\n"]
        visible_count = 0
        for d in debts:
            paid = await debt_paid_amount(session, d.id)
            remain = max(0, d.total_amount - paid)
            if remain <= 0:
                continue
            visible_count += 1
            turi = "📥 Qarz oldim" if getattr(d, 'debt_type', 'oldim') == 'oldim' else "📤 Qarz berdim"
            kim = "Kimdan" if getattr(d, 'debt_type', 'oldim') == 'oldim' else "Kimga"
            lines.append(
                f"#{d.id} — {turi}\n"
                f"{kim}: {getattr(d, 'person', '')}\n"
                f"Nomi: {d.name}\n"
                f"Jami: {money(d.total_amount)}\n"
                f"To'langan: {money(paid or 0)}\n"
                f"Qolgan: {money(remain)}\n"
                f"Izoh: {d.note}\n"
            )
        if visible_count == 0:
            return "Faol qarzlar yo'q. Barcha qarzlar to'liq yopilgan."
        return '\n'.join(lines)


@router.callback_query(F.data == 'debt_list')
async def debt_list(call: CallbackQuery):
    await call.message.answer(await debts_text())
    await call.answer()


@router.callback_query(F.data.in_({'debt_lend', 'debt_pay'}))
async def debt_pay_start(call: CallbackQuery, state: FSMContext):
    if not call.from_user or call.from_user.id != OWNER_ID:
        await call.answer("Bu amal siz uchun ruxsat etilmagan.", show_alert=True)
        return

    people = await active_borrowed_people()
    if not people:
        await call.message.answer("Hozir qarz beradigan odam yo'q. Faol qarzlaringiz mavjud emas.")
        await call.answer()
        return

    kb = InlineKeyboardBuilder()
    for index, item in enumerate(people):
        kb.button(
            text=f"👤 {item['person']} — {money(item['remaining'])}",
            callback_data=f"pay_person:{index}"
        )
    kb.adjust(1)
    await state.update_data(debt_people=people)
    await state.set_state(PayDebt.person)
    await call.message.answer(
        "Kimga qarz berdingiz?\n\nQarzingiz bor odamni tanlang:",
        reply_markup=kb.as_markup()
    )
    await call.answer()


@router.callback_query(F.data.startswith('pay_person:'))
async def pay_debt_pick(call: CallbackQuery, state: FSMContext):
    if not call.from_user or call.from_user.id != OWNER_ID:
        await call.answer("Bu amal siz uchun ruxsat etilmagan.", show_alert=True)
        return

    try:
        index = int(call.data.split(':', 1)[1])
    except (ValueError, IndexError):
        await call.answer("Noto'g'ri tanlov.", show_alert=True)
        return

    data = await state.get_data()
    people = data.get('debt_people', [])
    if index < 0 or index >= len(people):
        await call.answer("Ro'yxat eskirgan. Qaytadan tanlang.", show_alert=True)
        return

    item = people[index]
    await state.update_data(
        person_key=item['key'],
        person_name=item['person'],
        person_remaining=item['remaining'],
    )
    await state.set_state(PayDebt.amount)
    await call.message.answer(
        f"👤 Kimga: {item['person']}\n"
        f"💳 Hozirgi qarzingiz: {money(item['remaining'])}\n\n"
        "Qancha qarz berdingiz? Summani yozing.\n"
        "Masalan: 500000"
    )
    await call.answer()


@router.message(PayDebt.amount)
async def pay_debt_amount(message: Message, state: FSMContext):
    if not owner_only(message):
        return
    amount = parse_amount(message.text or '')
    if amount is None or amount <= 0:
        await message.answer("Summani faqat raqam bilan yozing.")
        return
    data = await state.get_data()
    person_key = data.get('person_key', '')
    person_name = data.get('person_name', "Noma'lum")

    async with SessionLocal() as session:
        debts = (
            await session.execute(
                select(Debt)
                .where(Debt.debt_type == 'oldim')
                .order_by(Debt.created_at.asc(), Debt.id.asc())
            )
        ).scalars().all()

        active_debts: list[tuple[Debt, int]] = []
        total_remaining = 0
        for debt in debts:
            debt_person_key = ((getattr(debt, 'person', '') or "Noma'lum").strip()).casefold()
            if debt_person_key != person_key:
                continue
            paid = await debt_paid_amount(session, debt.id)
            remaining = max(0, int(debt.total_amount) - paid)
            if remaining > 0:
                active_debts.append((debt, remaining))
                total_remaining += remaining

        if total_remaining <= 0:
            await state.clear()
            await message.answer(
                "Bu odamga qarzingiz allaqachon to'liq yopilgan.",
                reply_markup=main_menu()
            )
            return

        if amount > total_remaining:
            await message.answer(
                f"❌ Kiritilgan summa qarzdan ko'p.\n"
                f"Hozirgi qarz: {money(total_remaining)}\n\n"
                "Qarzdan oshmaydigan summani yozing."
            )
            return

        left_to_allocate = amount
        for debt, remaining in active_debts:
            if left_to_allocate <= 0:
                break
            payment_part = min(left_to_allocate, remaining)
            session.add(
                DebtPayment(
                    debt_id=debt.id,
                    amount=payment_part,
                    note=f"{person_name}ga qarz qaytarildi"
                )
            )
            left_to_allocate -= payment_part

        await session.commit()

    remaining_after = total_remaining - amount
    await state.clear()
    if remaining_after == 0:
        result_text = (
            f"✅ {person_name}ga qarz to'liq berildi.\n"
            f"To'langan summa: {money(amount)}\n\n"
            "Bu odam faol qarzlar ro'yxatidan olib tashlandi. "
            "Qarz tarixi Excel hisobotida saqlanib qoladi."
        )
    else:
        result_text = (
            f"✅ Qarzning bir qismi berildi.\n"
            f"Kimga: {person_name}\n"
            f"To'langan: {money(amount)}\n"
            f"Qolgan qarz: {money(remaining_after)}\n\n"
            "Qarz to'liq yopilmagani uchun ro'yxatda qoladi."
        )

    await message.answer(result_text, reply_markup=main_menu())




def autosize_sheets(wb: Workbook):
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = '' if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


async def create_excel_report(report_type: str = 'all', start: datetime | None = None, end: datetime | None = None) -> Path:
    Path('exports').mkdir(exist_ok=True)
    stamp = now_tashkent().strftime('%Y_%m_%d_%H_%M')
    path = Path('exports') / f"{report_type}_hisobot_{stamp}.xlsx"
    wb = Workbook()

    async with SessionLocal() as session:
        if start and end:
            incomes = (await session.execute(select(Income).where(Income.created_at >= start, Income.created_at <= end).order_by(Income.created_at.desc()))).scalars().all()
            expenses = (await session.execute(select(Expense).where(Expense.created_at >= start, Expense.created_at <= end).order_by(Expense.created_at.desc()))).scalars().all()
            pays = (await session.execute(select(DebtPayment).where(DebtPayment.created_at >= start, DebtPayment.created_at <= end).order_by(DebtPayment.created_at.desc()))).scalars().all()
        else:
            incomes = (await session.execute(select(Income).order_by(Income.created_at.desc()))).scalars().all()
            expenses = (await session.execute(select(Expense).order_by(Expense.created_at.desc()))).scalars().all()
            pays = (await session.execute(select(DebtPayment).order_by(DebtPayment.created_at.desc()))).scalars().all()
        debts = (await session.execute(select(Debt).order_by(Debt.created_at.desc()))).scalars().all()

        paid_by_debt = {}
        for d in debts:
            paid = await session.scalar(select(func.coalesce(func.sum(DebtPayment.amount), 0)).where(DebtPayment.debt_id == d.id))
            paid_by_debt[d.id] = int(paid or 0)

    ws = wb.active
    ws.title = 'Umumiy'
    income_total = sum(i.amount for i in incomes)
    expense_total = sum(e.amount for e in expenses)
    debt_pay_total = sum(p.amount for p in pays)
    ws.append(['Hisobot turi', report_type])
    ws.append(['Boshlanish', start.strftime('%Y-%m-%d %H:%M') if start else 'Barchasi'])
    ws.append(['Tugash', end.strftime('%Y-%m-%d %H:%M') if end else 'Barchasi'])
    ws.append(['Jami kirim', income_total])
    ws.append(['Jami chiqim', expense_total])
    ws.append(['Qolgan pul', income_total - expense_total])
    ws.append(["Qarz to'lovlari", debt_pay_total])

    ws1 = wb.create_sheet('Kirimlar')
    ws1.append(['Sana', 'Manba', 'Summa', 'Izoh'])
    for x in incomes:
        ws1.append([x.created_at.strftime('%Y-%m-%d %H:%M'), x.source, x.amount, x.note])

    ws2 = wb.create_sheet('Chiqimlar')
    ws2.append(['Sana', 'Kategoriya', 'Summa', 'Izoh'])
    for x in expenses:
        ws2.append([x.created_at.strftime('%Y-%m-%d %H:%M'), x.category, x.amount, x.note])

    ws3 = wb.create_sheet('Qarzlar')
    ws3.append(['ID', 'Sana', 'Turi', 'Kimdan/Kimga', 'Nomi', 'Jami summa', "To'langan", 'Qolgan', 'Izoh'])
    for x in debts:
        turi = 'Qarz oldim' if getattr(x, 'debt_type', 'oldim') == 'oldim' else 'Qarz berdim'
        paid = paid_by_debt.get(x.id, 0)
        ws3.append([x.id, x.created_at.strftime('%Y-%m-%d %H:%M'), turi, getattr(x, 'person', ''), x.name, x.total_amount, paid, x.total_amount - paid, x.note])

    ws4 = wb.create_sheet("Qarz to'lovlari")
    ws4.append(['Sana', 'Qarz ID', 'Kimga/Kimdan', 'Qarz nomi', 'Turi', 'Summa', 'Izoh'])
    debts_by_id = {d.id: d for d in debts}
    for x in pays:
        debt = debts_by_id.get(x.debt_id)
        debt_type = getattr(debt, 'debt_type', 'oldim') if debt else 'oldim'
        turi = 'Qarz qaytarildi' if debt_type == 'oldim' else 'Qarz qaytdi'
        ws4.append([
            x.created_at.strftime('%Y-%m-%d %H:%M'),
            x.debt_id,
            getattr(debt, 'person', '') if debt else '',
            getattr(debt, 'name', '') if debt else '',
            turi,
            x.amount,
            x.note
        ])

    autosize_sheets(wb)
    wb.save(path)
    return path


async def create_debts_excel() -> Path:
    Path('exports').mkdir(exist_ok=True)
    path = Path('exports') / f"qarzdorlik_{now_tashkent().strftime('%Y_%m_%d_%H_%M')}.xlsx"
    wb = Workbook()

    async with SessionLocal() as session:
        debts = (await session.execute(select(Debt).order_by(Debt.created_at.desc()))).scalars().all()
        payments = (
            await session.execute(
                select(DebtPayment).order_by(DebtPayment.created_at.desc())
            )
        ).scalars().all()

        paid_by_debt: dict[int, int] = {}
        for d in debts:
            paid_by_debt[d.id] = await debt_paid_amount(session, d.id)

    debts_by_id = {d.id: d for d in debts}

    ws_borrowed = wb.active
    ws_borrowed.title = 'Qarz olganlar'
    ws_borrowed.append([
        'ID', 'Qarz olingan sana', 'Kimdan', 'Qarz nomi', 'Olingan summa',
        'Qaytarilgan', 'Qolgan', 'Holati', 'Izoh'
    ])
    for d in debts:
        if getattr(d, 'debt_type', 'oldim') != 'oldim':
            continue
        paid = paid_by_debt.get(d.id, 0)
        remaining = max(0, int(d.total_amount) - paid)
        status = "To'liq yopilgan" if remaining == 0 else "Qarz mavjud"
        ws_borrowed.append([
            d.id,
            d.created_at.strftime('%Y-%m-%d %H:%M'),
            getattr(d, 'person', ''),
            d.name,
            d.total_amount,
            paid,
            remaining,
            status,
            d.note
        ])

    ws_lent = wb.create_sheet('Qarz berganlar')
    ws_lent.append([
        'ID', 'Qarz berilgan sana', 'Kimga', 'Qarz nomi', 'Berilgan summa',
        'Qaytgan', 'Qolgan', 'Holati', 'Izoh'
    ])
    for d in debts:
        if getattr(d, 'debt_type', 'oldim') != 'berdim':
            continue
        paid = paid_by_debt.get(d.id, 0)
        remaining = max(0, int(d.total_amount) - paid)
        status = "To'liq qaytgan" if remaining == 0 else "Qarz mavjud"
        ws_lent.append([
            d.id,
            d.created_at.strftime('%Y-%m-%d %H:%M'),
            getattr(d, 'person', ''),
            d.name,
            d.total_amount,
            paid,
            remaining,
            status,
            d.note
        ])

    ws_payments = wb.create_sheet('Qarz qaytarishlar')
    ws_payments.append([
        'To‘lov sanasi', 'Qarz ID', 'Kimga/Kimdan', 'Qarz nomi',
        'Amal turi', 'To‘langan summa', 'Izoh'
    ])
    for payment in payments:
        debt = debts_by_id.get(payment.debt_id)
        debt_type = getattr(debt, 'debt_type', 'oldim') if debt else 'oldim'
        action = 'Qarz berdim (qaytardim)' if debt_type == 'oldim' else 'Qarz qaytdi'
        ws_payments.append([
            payment.created_at.strftime('%Y-%m-%d %H:%M'),
            payment.debt_id,
            getattr(debt, 'person', '') if debt else '',
            getattr(debt, 'name', '') if debt else '',
            action,
            payment.amount,
            payment.note
        ])

    autosize_sheets(wb)
    wb.save(path)
    return path


async def auto_excel_sender(bot: Bot):
    last_daily = None
    last_monthly = None
    while True:
        try:
            now = now_tashkent()
            if now.hour == 23 and now.minute >= 55 and last_daily != now.date():
                start = datetime(now.year, now.month, now.day, 0, 0, 0)
                end = datetime(now.year, now.month, now.day, 23, 59, 59)
                path = await create_excel_report('kunlik', start, end)
                await bot.send_document(OWNER_ID, FSInputFile(path), caption=f"✅ Kunlik Excel hisobot: {now.strftime('%Y-%m-%d')}")
                last_daily = now.date()

            last_day = calendar.monthrange(now.year, now.month)[1]
            month_key = f"{now.year}-{now.month}"
            if now.day == last_day and now.hour == 23 and now.minute >= 50 and last_monthly != month_key:
                start = datetime(now.year, now.month, 1, 0, 0, 0)
                end = datetime(now.year, now.month, last_day, 23, 59, 59)
                path = await create_excel_report('oylik', start, end)
                await bot.send_document(OWNER_ID, FSInputFile(path), caption=f"✅ Oylik Excel hisobot: {now.strftime('%Y-%m')}")
                last_monthly = month_key
        except Exception as e:
            print('Avtomatik Excel yuborishda xato:', e)
        await asyncio.sleep(60)


@router.callback_query(F.data == 'debt_excel')
async def debt_excel(call: CallbackQuery):
    path = await create_debts_excel()
    await call.message.answer_document(FSInputFile(path), caption='✅ Qarzdorlik Excel tayyor')
    await call.answer()


@router.message(F.text == "📊 Oylik hisobot")
async def monthly_report(message: Message):
    if not owner_only(message): return
    now = now_tashkent()
    start = datetime(now.year, now.month, 1)
    end = datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1], 23, 59, 59)
    async with SessionLocal() as session:
        income_total = await session.scalar(select(func.coalesce(func.sum(Income.amount), 0)).where(Income.created_at >= start, Income.created_at <= end))
        expense_total = await session.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(Expense.created_at >= start, Expense.created_at <= end))
        cats = (await session.execute(select(Expense.category, func.sum(Expense.amount)).where(Expense.created_at >= start, Expense.created_at <= end).group_by(Expense.category).order_by(func.sum(Expense.amount).desc()))).all()
        debt_paid = await session.scalar(select(func.coalesce(func.sum(DebtPayment.amount), 0)).where(DebtPayment.created_at >= start, DebtPayment.created_at <= end))
    lines = [
        f"📊 {now.strftime('%B %Y')} hisoboti",
        f"\nKirim: {money(income_total or 0)}",
        f"Chiqim: {money(expense_total or 0)}",
        f"Qolgan: {money((income_total or 0) - (expense_total or 0))}",
        f"Qarz to'lovlari: {money(debt_paid or 0)}",
        "\nKategoriya bo'yicha chiqimlar:"
    ]
    if cats:
        for cat, total in cats:
            lines.append(f"- {cat}: {money(total)}")
    else:
        lines.append("- Hali chiqim yo'q")
    await message.answer('\n'.join(lines))


@router.message(F.text == "📤 Excel eksport")
async def excel_export(message: Message):
    if not owner_only(message): return
    path = await create_excel_report('umumiy')
    await message.answer_document(FSInputFile(path), caption="✅ Umumiy Excel hisobot tayyor")


@router.message(F.text == "🔍 Qidirish")
async def search_start(message: Message, state: FSMContext):
    if not owner_only(message): return
    await state.set_state(SearchState.query)
    await message.answer("Qidiriladigan so'zni yozing. Masalan: gaz, mashina, oylik")


@router.message(SearchState.query)
async def search_query(message: Message, state: FSMContext):
    q = f"%{message.text}%"
    async with SessionLocal() as session:
        incomes = (await session.execute(select(Income).where(Income.note.ilike(q) | Income.source.ilike(q)).limit(10))).scalars().all()
        expenses = (await session.execute(select(Expense).where(Expense.note.ilike(q) | Expense.category.ilike(q)).limit(10))).scalars().all()
    lines = ["🔍 Qidiruv natijalari:\n"]
    for x in incomes:
        lines.append(f"➕ {x.source}: {money(x.amount)} — {x.note}")
    for x in expenses:
        lines.append(f"➖ {x.category}: {money(x.amount)} — {x.note}")
    if len(lines) == 1:
        lines.append("Hech narsa topilmadi.")
    await state.clear()
    await message.answer('\n'.join(lines), reply_markup=main_menu())


@router.message()
async def unknown(message: Message):
    if not owner_only(message):
        await message.answer("⛔ Bu bot faqat egasi uchun ishlaydi.")
        return
    await message.answer("Pastdagi tugmalardan foydalaning.", reply_markup=main_menu())


async def main():
    await init_db()
    bot = Bot(BOT_TOKEN)
    # Boshqa deploy/webhook/pending update muammolarini tozalaydi
    await bot.delete_webhook(drop_pending_updates=True)
    asyncio.create_task(auto_excel_sender(bot))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    print('✅ Moliyaviy nazorat bot ishga tushdi...')
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == '__main__':
    asyncio.run(main())
